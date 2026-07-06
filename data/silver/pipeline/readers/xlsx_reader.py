import hashlib
import io
import os

from data.silver.pipeline.readers.base import BaseReader
from data.silver.schemas.source import SourceData


class XlsxReader(BaseReader):
    def can_handle(self, source: SourceData) -> bool:
        ext = os.path.splitext(source.filename.lower())[1]
        return ext in (".xlsx", ".xls")

    def read(self, source: SourceData) -> dict:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(source.raw_data), read_only=True, data_only=True)
            lines = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                lines.append(f"=== Sheet: {sheet_name} ===")
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    lines.append(" | ".join(cells))
            content = "\n".join(lines)
        except Exception:
            content = ""
        return {"content": content}

    def extract_metadata(self, source: SourceData) -> dict:
        ext = os.path.splitext(source.filename.lower())[1]
        checksum = hashlib.sha256(source.raw_data).hexdigest()
        return {
            "extension": ext,
            "checksum": checksum,
            "size_bytes": source.size_bytes,
            "object_key": source.object_key,
            "bucket": source.bucket,
            "source": f"minio://{source.bucket}/{source.object_key}",
            "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }
